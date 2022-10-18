from lib2to3.pgen2 import driver
from tokenize import String
from unicodedata import name
from webbrowser import Chrome
import undetected_chromedriver.v2 as uc
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.keys import Keys
import time
import os
import xlsxwriter
import re
from datetime import date
import pandas as pd
from openpyxl import load_workbook
import openpyxl
import pyautogui
import pyperclip
from selenium.webdriver import ActionChains







today = str(date.today())
filename = "ig_list.xlsx"
wb = load_workbook(filename)
ws = wb.worksheets[0]
row = str(int(len(ws["A"]))+1)
ws["A" + row] = today






# Mon 手工禮物 acc: beefrice password: deliveryape
# Tue 香薰蠟燭 acc: friedfishnchips password: deliveryape
# Wed 永生花 acc: dot._.dot.dot password: deliveryape
# Thurs 寵物用品 acc: wtda.faq password: deliveryape
#Fri 水晶 acc: delivery_ape_emma password: deliveryape@







if __name__ ==  '__main__':

   class My_Chrome(uc.Chrome):
    def __del__(self):
        pass

    options = webdriver.ChromeOptions()
    driver = uc.Chrome(
        options=options,
    )

    options.add_argument("--user-data-dir=/private/var/folders/qz/y1ddhm2d5s194mj9cs31vqfr0000gn/T/.com.google.Chrome.3APWJG/Default")
    options.add_argument('--no-first-run --no-service-autorun --password-store=basic')




    driver.get("https://www.instagram.com/")
    array = []
    link_array = []
    count = 0
    workbook = xlsxwriter.Workbook(today + ".xlsx")
    worksheet = workbook.add_worksheet()
    





    username = WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.NAME, "username"))
    )
    password = WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.NAME, "password"))
    )
    login = driver.find_element(By.XPATH,'//*[@id="loginForm"]/div/div[3]/button')

    username.clear()
    password.clear()
    username.send_keys('delivery_ape_emma')
    password.send_keys('deliveryape@')
    login.click()

    #click search bar

    search = WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.XPATH, '//*[@id="react-root"]/section/nav/div[2]/div/div/div[2]/input'))
    )

    webdriver.ActionChains(driver).move_to_element(search ).click(search ).perform()
    time.sleep(1)
    webdriver.ActionChains(driver).move_to_element(search ).click(search ).perform()

    # click the search bar recoprd

    click = WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.XPATH, '//*[@id="react-root"]/section/nav/div[2]/div/div/div[2]/div[3]/div/div[2]/div/ul/div/a/div'))
    )

    webdriver.ActionChains(driver).move_to_element(click ).click(click ).perform()


    #wait the result load

    time.sleep(5)

    num_list=[]

    #number of loops
    num = WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.CLASS_NAME, '_ac2a'))
    )

    #num = num.get_attribute("innerHTML")
    
    #num1 = re.sub(",","",num)

    #num1 = int(num1)

    

    lists = []

    list2 =[]

    list3 = []

    

    

     

    for times in range (1):

        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")

        time.sleep(5)


        appear =  driver.find_elements(By.CSS_SELECTOR, '.qi72231t.nu7423ey.n3hqoq4p.r86q59rh.b3qcqh3k.fq87ekyn.bdao358l.fsf7x5fv.rse6dlih.s5oniofx.m8h3af8h.l7ghb35v.kjdc1dyq.kmwttqpk.srn514ro.oxkhqvkx.rl78xhln.nch0832m.cr00lzj9.rn8ck1ys.s3jn8y49.icdlwmnq._a6hd')

    

        for _ in appear:
            link = _.get_attribute("href")
            lists.append(link)

        for b in lists:
            if b != 'https://www.instagram.com/' :
                list2.append(b)
        
        for a in list2:
            if a == 'https://www.instagram.com/direct/inbox/':
                list2.remove(a)

        for a in list2:
            if a =='https://www.instagram.com/explore/':
                list2.remove(a)

        for kan in list2:
            if kan not in list3:
                list3.append(kan)

    


    print("this is a line")
    print("this is a line")
    print("this is a line")
    print(list3)

    for url in list3:
        

        
                    
        driver.switch_to.new_window()
        driver.get(url)
        

                     
        shop_name = WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.CSS_SELECTOR, '.qi72231t.nu7423ey.n3hqoq4p.r86q59rh.b3qcqh3k.fq87ekyn.bdao358l.fsf7x5fv.rse6dlih.s5oniofx.m8h3af8h.l7ghb35v.kjdc1dyq.kmwttqpk.srn514ro.oxkhqvkx.rl78xhln.nch0832m.cr00lzj9.rn8ck1ys.s3jn8y49.icdlwmnq._acan._acao._acat._acaw._a6hd'))) 
        print("sucess")
        shop_page = shop_name.get_attribute("href")
        driver.switch_to.new_window()
        driver.get(shop_page)
        shop_name = WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.CSS_SELECTOR, "._aacl._aacs._aact._aacx._aada"))) 
        shop_name_text = shop_name.get_attribute("innerHTML")
        bio = driver.find_elements(By.CSS_SELECTOR,'._aacl._aacp._aacu._aacx._aad6._aade')
        shop_info =[] #put all items in bio in a list
        for _ in bio:
            bio_text = _.get_attribute("innerText")
            shop_info.append(bio_text)

        followers = shop_info[1] 
        bio_text = shop_info[3]

        #determine delivery method in bio

        if "順"  in bio_text or "SF" in bio_text:
            delivery = "SF"

        elif "面交" in bio_text or "地鐵站" in bio_text or "交收" in bio_text:
            delivery = "面交"
    
        elif "郵" in bio_text or "Post" in bio_text:
            delivery = "平郵"

        elif  "蝦皮" in bio_text or "私訊" in bio_text or "Line" in bio_text or "line" in bio_text or "台灣" in bio_text or "露天" in bio_text or "台北" in bio_text or "高雄" in bio_text:
            delivery = "台灣"

        else:
            delivery = "unknown"
        
        df = pd.read_excel (r'/Users/wailokkwok/Desktop/fuck/ig_list.xlsx', sheet_name='Sheet 1')
        database = df['ig'].tolist()

        
        if shop_name_text not in database:     
            row = str(int(len(ws['A']))+1)
            ws["A" + row] = shop_name_text               
            ws["G" + row] = shop_page
            ws["B" +row ] = followers
            ws["C" +row ] = delivery
            wb.save('/Users/wailokkwok/Desktop/fuck/ig_list.xlsx')

            if delivery != "台灣":
                follow = WebDriverWait(driver, 10).until(
                    EC.presence_of_element_located((By.CSS_SELECTOR, '._aacl._aaco._aacw._adda._aad6._aade'))
                    )

                follow.click()

                time.sleep(2)

                chat = WebDriverWait(driver, 10).until(
                    EC.presence_of_element_located((By.CSS_SELECTOR, '._acan._acap._acat'))
                    )

                chat.click()

                time.sleep(5)

                textbox = driver.find_element(By.CSS_SELECTOR,'._ab8w._ab94._ab99._ab9f._ab9m._ab9o._abbh._abcm')
                textbox.click

                print("c")

                time.sleep(2)
                textbox.click
                time.sleep(1)
                textbox.click
                pyperclip.copy("""Hello🙌🏻🙌🏻我哋係一間大學生組成嘅公司，專幫Carousell 用戶解決面交問題🙈🙈

        #我哋係地鐵站附近擺左智能櫃，你哋擺完啲貨入去就可以叫個客自己拎，唔駛再約時間🤣目前旺角同銅鑼灣兩個地方都有智能櫃！🤓

        #如果有興趣，可以到Instagram/網站下單！😋

        #Instagram：delivery_ape
        #WhatsApp：5631 6150""")

                act = ActionChains(driver)
                act.key_down(Keys.META)
                act.send_keys('v')
                act.key_up(Keys.META)
                act.perform() 
                time.sleep(3)
                pyautogui.press('return')



                aaa = driver.find_elements(By.XPATH,"//*[name()='svg']")
                
                aaa[-2].click()

                time.sleep(2)
        
                pyautogui.write('/Users/wailokkwok/Desktop/fuck/a.jpg',interval=0.15)
                time.sleep(2)
                pyautogui.press('return')
                time.sleep(2)
                pyautogui.press('return')
                time.sleep(2)
                pyautogui.press('return')

                time.sleep(2)  
                driver.close()
                driver.switch_to.window(driver.window_handles[-1])
                driver.close()
                driver.switch_to.window(driver.window_handles[-1])
                time.sleep(10)

            else:
                driver.close()
                driver.switch_to.window(driver.window_handles[-1])
                driver.close()
                driver.switch_to.window(driver.window_handles[-1])
                time.sleep(5)
                

        else:
            driver.close()
            driver.switch_to.window(driver.window_handles[-1])
            driver.close()
            driver.switch_to.window(driver.window_handles[-1])
            time.sleep(5)

    

        print(database)

       
        


      

    
    

    
    
    


   

    

       
            
    


    


   
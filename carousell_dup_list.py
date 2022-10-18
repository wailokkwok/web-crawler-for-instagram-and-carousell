from tokenize import String
from unicodedata import name
from webbrowser import Chrome
from torch import amax
import undetected_chromedriver.v2 as uc
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.keys import Keys
from selenium.webdriver import ActionChains
from selenium import webdriver
import pyperclip
import time
import os
import xlsxwriter
from datetime import date
today = str(date.today())
import pyautogui
import pandas as pd
from openpyxl import load_workbook
import openpyxl


#pip install pyperclip
#pip install pyautogui
#pip install pandas
#pip install xlrd
#pip install openpyxl

today = str(date.today())
filename = "carousell_list.xlsx"
wb = load_workbook(filename)
ws = wb.worksheets[0]
row = str(int(len(ws["A"]))+1)
ws["A" + row] = today







if __name__ ==  '__main__':
    

   class My_Chrome(uc.Chrome):
    def __del__(self):
        pass

    options = webdriver.ChromeOptions()
    browser = uc.Chrome(
        options=options,
        executable_path=r'C:\WebDrivers\chromedriver.exe'
    )

    options.add_argument("--user-data-dir=/private/var/folders/qz/y1ddhm2d5s194mj9cs31vqfr0000gn/T/.com.google.Chrome.3APWJG/Default")
    options.add_argument('--no-first-run --no-service-autorun --password-store=basic')
    options.add_experimental_option("detach", True)
    


    lists = []
    array = []
    count = 0
    workbook = xlsxwriter.Workbook(today + "Caoursell" + ".xlsx")
    worksheet = workbook.add_worksheet()
    
   





    

    browser.get("https://www.carousell.com.hk/login/")
    username = WebDriverWait(browser, 1000).until(
        EC.presence_of_element_located((By.NAME, "username"))
    )
    password = WebDriverWait(browser, 1000).until(
        EC.presence_of_element_located((By.NAME, "password"))
    )

    login = WebDriverWait(browser, 1000).until(
        EC.presence_of_element_located((By.XPATH, '//*[@id="root"]/div[2]/form/button'))
    )

    username.clear()
    password.clear()
    username.send_keys('delivery_ape_trey')
    password.send_keys('deliveryape')
    login.click()

    search = WebDriverWait(browser, 1000).until(
        EC.presence_of_element_located((By.XPATH, '//*[@id="root"]/div/div[1]/div[1]/div/div[2]/div[1]/div/form/div[1]/div/div/div/input'))
    )

    search.click()

    search.send_keys("手工")
    time.sleep(2)

    search.send_keys(Keys.RETURN)
    time.sleep(2)


    for times in range (6):
        
        browser.execute_script("window.scrollTo(0, document.body.scrollHeight);")

        time.sleep(2)

        more = WebDriverWait(browser, 1000).until(
        EC.presence_of_element_located((By.XPATH, '//*[@id="root"]/div/div[3]/div/div[5]/main/div/button')))


        time.sleep(2)
        more.click()
    
    a_tag = browser.find_elements(By.TAG_NAME, "a")

    aax =[]

    



    for _ in a_tag:
        href =_.get_attribute("href")
        if "tap_index=" in href:
            aax.append(href)

    for aay in aax:
        browser.switch_to.new_window()
        browser.get(aay)
        shop_name_tag = WebDriverWait(browser, 10).until(
        EC.presence_of_element_located((By.XPATH, '//*[@id="root"]/div/div[3]/div[2]/div[2]/div[1]/div[1]/div[1]/a/span')))
        shop_name = shop_name_tag.get_attribute("innerText")
        print(shop_name)
        df = pd.read_excel (r'/Users/wailokkwok/Desktop/fuck/carousell_list.xlsx', sheet_name='Sheet1')
        database = df['user name'].tolist()
 
        
        if shop_name not in database:
            
            
            row = str(int(len(ws['A']))+1)
            ws["A" + row] = shop_name               
            wb.save('/Users/wailokkwok/Desktop/fuck/carousell_list.xlsx')
            print(shop_name)
            chat = browser.find_element(By.XPATH,'//*[@id="root"]/div/div[3]/div[2]/div[2]/div[1]/div[2]/button')
            print("hi")
            chat.click()
            time.sleep(5)
            box = browser.find_element(By.XPATH,'//*[@id="root"]/div/div[2]/div/div[2]/div[4]/div[1]/div[1]/textarea')
            box.click()
            print("hi")
            time.sleep(2)
            pyperclip.copy("""Hello🙌🏻🙌🏻我哋係一間大學生組成嘅公司，專幫Carousell 用戶解決面交問題🙈🙈

#我哋係地鐵站附近擺左智能櫃，你哋擺完啲貨入去就可以叫個客自己拎，唔駛再約時間🤣目前旺角同銅鑼灣兩個地方都有智能櫃！🤓

#如果有興趣，可以到Instagram/網站下單！😋

#Instagram：delivery_ape
#WhatsApp：5631 6150""")

            act = ActionChains(browser)
            act.key_down(Keys.META)
            act.send_keys('v')
            act.key_up(Keys.META)
            act.perform() 
            box.send_keys(Keys.RETURN)

            image = WebDriverWait(browser, 30).until(
            EC.presence_of_element_located((By.XPATH, '//*[@id="root"]/div/div[2]/div/div[2]/div[4]/div[2]/label/div')))
            print("hi")
            image.click()
            time.sleep(2)
            pyautogui.write('/Users/wailokkwok/Desktop/fuck/a.jpg',interval=0.15)
            time.sleep(2)
            pyautogui.press('return')
            time.sleep(2)
            pyautogui.press('return')
            time.sleep(2)
            pyautogui.press('return')
            time.sleep(10)
            browser.close()
            browser.switch_to.window(browser.window_handles[-1])

        else:
            browser.close()
            browser.switch_to.window(browser.window_handles[-1])

    

 


            

   

        
            

        
       





    



    





    
            

            

                
                         
    
    

   









   





